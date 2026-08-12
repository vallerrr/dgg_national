"""
# Created by valler at 11/08/2026
Feature: rebuilds the UN World Population Prospects age/sex panel that the rest of the pipeline
reads as `un_1950_2023_processed.csv` (params.RAW), from the raw WPP single-age workbooks.

The script that originally produced `un_1950_2023_processed.csv` was never committed to any of the
three repos — only its consumers survive (13_coherent_ggi.py, 15_, 16_, 07_facebook_data_monthly.py).
This reconstructs it from the published file's own schema: 27 age bands x {both sexes, female,
male} plus the female/male ratio, in the exact column order of the shipped CSV, so the output is a
drop-in replacement.

`01_population_data.R` is the surviving sibling of that lost script: same 27 bands, same column
offsets, but keyed on `survey_start`, restricted to Year > 2001, and without the `_t` (both-sexes)
columns. The band definitions below are transcribed from it.

Input : WPP single-age workbooks (BOTH_SEXES / MALE / FEMALE), sheet "Estimates"
Output: params.UN_POP / un_1950_2023_processed_wpp2024_<date>.csv

The "Estimates" sheet stops at 2023. The file currently in params.RAW runs to 2024 because its
last year came from a projection sheet; nothing here reads a projection, so predictions for 2024+
(e.g. the Rwanda 2025 survey) have no population row until "Medium variant" is added too.

Run from the project root:  python src/01_01_un_population_wpp.py
"""
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import params

# The workbooks carry a preamble of blank and citation rows before the real header, and its length
# differs between them — BOTH_SEXES has four more leading blank rows than MALE/FEMALE. R's readxl
# trims leading blanks, so a hardcoded "row 13" worked there by accident; here the header is
# located by content instead.
HEADER_KEY = 'ISO3 Alpha-code'
# Columns 1:11 are metadata, columns 12:112 are single ages 0..100+, so column index = 12 + age.
AGE0_COL = 12
LAST_AGE_COL = 112  # age "100+"

# Age bands, in the order the shipped CSV lists them. Values are the 1-based *column* ranges used
# by 01_population_data.R's `c_across()`, kept as-is so the bands stay byte-compatible with the
# published file. Ranges are inclusive of both endpoints, and so are the bands themselves:
# "14_15" is ages 14 and 15, not a 14-to-15 interval.
BANDS = {
    '18_inf': (30, 112), '14_15': (26, 27), '15_16': (27, 28), '16_17': (28, 29),
    '17_18':  (29, 30),  '18_19': (30, 31), '13_14': (25, 26), '15_19': (27, 31),
    '20_24':  (32, 36),  '25_29': (37, 41), '30_34': (42, 46), '35_39': (47, 51),
    '40_44':  (52, 56),  '45_49': (57, 61), '50_54': (62, 66), '55_59': (67, 71),
    '60_64':  (72, 76),  '18_23': (30, 35), '20_inf': (32, 112), '20_64': (32, 76),
    '21_inf': (33, 112), '25_inf': (37, 112), '25_49': (37, 61), '25_64': (37, 76),
    '50_inf': (62, 112), '60_inf': (72, 112), '65_inf': (77, 112),
}

YEAR_MIN, YEAR_MAX = 1950, 2023
THOUSANDS = 1000  # WPP publishes thousands; the pipeline works in persons


def read_wpp_bands(path, suffix, sheet='Estimates'):
    """Read one WPP single-age workbook and collapse it to the 27 age bands.

    Streamed row by row: the workbooks are ~220MB and only the "Estimates" sheet is wanted, so
    materialising the whole thing costs minutes and gigabytes for no gain.

    Args:
        path: workbook path.
        suffix: 'f', 'm' or 't' — appended to each band name.
        sheet: worksheet to read; 'Estimates' is 1950-2023, projections live on other sheets.

    Returns:
        DataFrame with iso3, Year and one column per band, in persons.
    """
    print(f'reading {path.name} ...', flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    # These workbooks declare a <dimension> that covers only the preamble block; in read_only mode
    # openpyxl believes it and would stop after ~18 rows and 12 columns.
    ws.reset_dimensions()

    header, records = None, []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if HEADER_KEY in row:
                header = {name: j for j, name in enumerate(row) if name is not None}
                if len(row) != LAST_AGE_COL:
                    raise ValueError(f'{path.name}: expected {LAST_AGE_COL} columns, got {len(row)}')
            continue
        if row[header['Type']] != 'Country/Area':
            continue
        year = row[header['Year']]
        if year is None or not (YEAR_MIN <= int(year) <= YEAR_MAX):
            continue
        records.append((row[header['ISO3 Alpha-code']], int(year),
                        row[AGE0_COL - 1:LAST_AGE_COL]))
    wb.close()

    iso3 = [r[0] for r in records]
    years = [r[1] for r in records]
    # Suppressed cells come through as '...'; to_numeric turns those into NaN, as as.numeric does.
    ages = pd.DataFrame([r[2] for r in records]).apply(pd.to_numeric, errors='coerce').to_numpy()

    out = pd.DataFrame({'iso3': iso3, 'Year': years})
    for band, (first, last) in BANDS.items():
        # workbook column index -> 0-based index within `ages`
        cols = slice(first - AGE0_COL, last - AGE0_COL + 1)
        out[f'{band}_{suffix}'] = ages[:, cols].sum(axis=1) * THOUSANDS
    return out


def build_panel(edition='WPP2024'):
    """Assemble the three workbooks into the shipped 110-column schema."""
    def wpp(n, sex):
        return params.UN_POP_RAW / f'{edition}_POP_F01_{n}_POPULATION_SINGLE_AGE_{sex}.xlsx'

    pop_t = read_wpp_bands(wpp(1, 'BOTH_SEXES'), 't')
    pop_m = read_wpp_bands(wpp(2, 'MALE'), 'm')
    pop_f = read_wpp_bands(wpp(3, 'FEMALE'), 'f')

    panel = pop_t.merge(pop_f, on=['iso3', 'Year'], how='outer') \
                 .merge(pop_m, on=['iso3', 'Year'], how='outer')

    # `_r` is the female/male ratio, one per band, appended after the level columns.
    for band in BANDS:
        panel[f'{band}_r'] = panel[f'{band}_f'] / panel[f'{band}_m']

    # Column order of the shipped file: iso3, Year, then (t, f, m) interleaved per band, then _r.
    col_order = (['iso3', 'Year']
                 + [f'{b}_{s}' for b in BANDS for s in ('t', 'f', 'm')]
                 + [f'{b}_r' for b in BANDS])
    return panel[col_order]


if __name__ == '__main__':
    edition = 'WPP2024'
    panel = build_panel(edition)

    stamp = date.today().strftime('%Y%m%d')
    out_file = params.UN_POP / f'un_1950_2023_processed_{edition.lower()}_{stamp}.csv'
    out_file.parent.mkdir(parents=True, exist_ok=True)
    panel.to_csv(out_file, index=False)

    print(f'wrote {out_file}  ({len(panel)} rows, {panel.shape[1]} cols, '
          f'{panel.iso3.nunique()} countries, {panel.Year.min()}-{panel.Year.max()})')
